import os, subprocess,pymysql
from PyQt5 import uic # QtCore,
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QDate #, Qt
import configparser

config = configparser.ConfigParser()
config.read(r"./register/config.ini")
Denomination_name = config['Denomination_name']['denomination']
church_name_s = config['Church_name']['name']
church_addr = config['Church_addr']['address']
church_biz_no = config['biz_no']['biz_No']
church_name = f"{Denomination_name} {church_name_s}"

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

cur_fold = os.getcwd()
Location = os.path.join(cur_fold, "cont_issue") # 챗GPT 수정

saved_file = os.path.join(cur_fold, "DB","contribution_list.xlsx") # 책GPT 수정
imsi_file= os.path.join(cur_fold, "Tmp","imsi_file.xlsx")

today = QDate.currentDate()
form_class = uic.loadUiType(os.path.join(cur_fold,"ui","contribution_form.ui"))[0]
imsi_data=[]

class contribution_issue_go(QDialog, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        
        self.setWindowIcon(QIcon(os.path.join(cur_fold,'img','logo.ico')))
        self.setWindowTitle('기부금영수증 발급')

        global target_de_year,issue_year, church_name # issue_No  
        self.issue_date = QDateEdit()
        self.issue_date_widget.setDate(today)
        issue_year = str(today.year())
        self.jujung_ratio_widget.setText('10')
        target_de_year = str(today.year()-1)
        self.category_widget.setText('종교단체')
        self.issue_name_widget.setText("성도명")
        self.target_year_widget.setText(target_de_year)
        self.add_re_issue_focus = False
        church_name_text = church_name
        self.issue_church_widget.setText(church_name_text)
        self.issue_name_widget.setFocus()

        self.issue_name_widget.editingFinished.connect(self.issue_name_code)
        # self.target_year_widget.cursorPositionChanged.connect(self.issue_target_year)
        self.target_year_widget.editingFinished.connect(self.issue_target_year)
        self.issue_amount_widget.editingFinished.connect(self.on_changed_item)
        self.target_year_widget.editingFinished.connect(self.add_or_re_issue)
        self.issue_amount_widget.editingFinished.connect(self.add_or_re_issue)
        self.jumin_no_widget.editingFinished.connect(self.add_or_re_issue)
        self.address_widget.editingFinished.connect(self.add_or_re_issue)
        

    def button_coll(self):
        issue_Button = QPushButton("발급하기")
        issue_Button.clicked.connect(self.contribution_issue_s)
        close_Button = QPushButton("종료")
        close_Button.clicked.connect(self.issue_close)
    
    def showEvent(self, event):
        super(contribution_issue_go, self).showEvent(event)
        self.set_cursor_position()

    def set_cursor_position(self):
        # 커서를 특정 위치로 설정합니다 (예: QLineEdit)
        self.issue_name_widget.setText("성도명")
        self.issue_name_widget.setFocus()
        self.issue_name_widget.setCursorPosition(0)  # 커서를 맨 앞에 위치시킵니다
    
    def on_changed_item(self):  # 테이블의 숫자값이 변경되고 ' , '  넣어주기
        amount_text = self.issue_amount_widget.text()
        if amount_text.strip():
            try:
                value = int(amount_text.replace(",", ""))
                self.issue_amount_widget.setText(f"{value:,}")  # 숫자를 쉼표로 포맷팅
            except ValueError:
                QMessageBox.about(None,'입력오류',"올바르지 않은 숫자입니다.")
    def add_or_re_issue(self):
        self.add_re_issue_focus = True

    def issue_name_code(self):
        from basic.sungdo import s_name_select, name_code_select
        global  name_diff, hap_code, r_name
        if self.add_re_issue_focus == True:
            self.target_year_widget.clear()
            self.issue_name_hapcode_widget.clear()
            self.issue_amount_widget.clear()
            self.issable_amount.clear()
            self.total_amount_widget.clear()

        Na_code = name_code_select() # 코드로 쓰이는 개별성명 선택
        name_diff = self.issue_name_widget.text()

        # 시작일과 종료일
        
        if name_diff != "성도명" and name_diff: 
            if (name_diff,) in Na_code : 
                codename = s_name_select(name_diff)
                hap_code = codename[0][0]; r_name = codename[0][1]  # 발행대상자가 1명 이므로 iloc[0, ]에서 행은 항상 0 이다.
                hap_code = hap_code.strip("[',']")
                self.issue_name_hapcode_widget.setText(hap_code)
            else: #except IndexError: 
                QMessageBox.critical(self, "검색에러","등록되지 않은 이름입니다.!!!")
                self.issue_name_hapcode_widget.clear()
                self.issue_name_widget.setFocus()
                return       
        else: #except IndexError: 
            if name_diff != "성도명":
                QMessageBox.critical(self, "입력오류","'성도명'을 입력하여 주세요.!!!")
                # QMessageBox.critical(None, "검색내용에러","등록되지 않은 이름입니다.!!!")
                self.issue_name_hapcode_widget.clear()
                self.issue_name_widget.setFocus()
                return
            
    # def confirm_year(self,target1_year):  # 테이블의 숫자값이 변경되고 ' , '  넣어주기
    #     try:
    #         if target1_year != "" and target1_year != "-" :
    #             if int(target1_year.replace(",", "")) > 2000 and int(target1_year.replace(",", "")) < 10000 :
    #                 pass
    #             else:
    #                 QMessageBox.about(None,'입력오류',"'년도'가 올바르지 않습니다.")
    #     except ValueError:
    #         QMessageBox.about(None,'입력오류',"'년도'가 올바르지 않습니다.")
    #         self.target_year_widget.clear()
    #         self.target_year_widget.setFocus()
    #         return

    def issue_target_year(self):
        from basic.contribution_issue import issued_count, issued_amount_hap
        from basic.hun_split import contribution_amount
        global  issue_No, tar_year, issuable_amount, targ_year, i_amount, jujung_ratio, sv_date, ev_date
        hap_code = self.issue_name_hapcode_widget.text()
        target1_year = self.target_year_widget.text()
        if name_diff != "성도명":
            if target1_year:
                s_date = target1_year + '-01-01'
                if len(s_date) == 9:
                    s_date = s_date[:5] + '0' + s_date[5:]
                e_date = target1_year + '-12-31'
                sv_date = datetime.strptime(s_date,"%Y-%m-%d")
                ev_date = datetime.strptime(e_date,"%Y-%m-%d")
                Y1 = today.year()
                is_count = str(issued_count(Y1)).zfill(3)          #발행번호
                issue_No = '일산백석'+str(issue_year)[2:]+'-'+ is_count       # 발행기호 +발행번호)
                self.issue_no_widget.setText(issue_No)

                #try:
                if hap_code != '' or  hap_code != None :
                    self.category_widget.setText('종교단체')
                    self.issue_name_hapcode_widget.setText(hap_code)
                    tar_year = int(target1_year)
                    targ_year = str(tar_year)
                    jujung_ratio = int(self.jujung_ratio_widget.text())/100
                    to_hap = int(round(int(contribution_amount(tar_year,hap_code))*(1+jujung_ratio), -3)) # 발행할 수 있는 총액 구하기(천 단위 이하 반올림)
                    total_hap = format(to_hap,',')
                    
                    self.total_amount_widget.setText(total_hap)
                    issued_amount = issued_amount_hap(tar_year,hap_code)   # 발행된 금액
                    issuable_amount = to_hap - issued_amount                 # 발행 가능 금액
                    i_amount = format(issuable_amount,",")
                    self.issable_amount.setText(i_amount)
                    self.issue_amount_widget.setText(i_amount)        # 발행가능금액 텍스트 표기
                    
                else :
                    QMessageBox.critical(self, '검색내용에러', "등록되지 않은 이름입니다.!!!")
                    # QMessageBox.about(self,'입력오류 !!!','등록되지 않은 이름 입니다. 확인하여 주십시오')
                    self.issue_name_widgetsetFocus()
                    return
                    # self.target_year_widget.clear()

            else:
                QMessageBox.about(self,'입력오류 !!!','발행할 대상년도가 없습니다.')
                self.issue_name_widget.clear()
                self.issue_name_hapcode_widget.clear()
                self.target_year_widget.clear()
                hap_code = ""
        else:
            QMessageBox.critical(self, "입력오류","'성도명'을 입력하여 주세요.!!!")
            self.issue_name_hapcode_widget.clear()
            self.issue_name_widget.setFocus()
            return
        
    def close_reset(self):
        self.issue_name_widget.clear()
        self.target_year_widget.clear()
        self.issue_amount_widget.clear()
        self.total_amount_widget.clear()
        self.issue_name_hapcode_widget.clear()
        self.issable_amount.clear()
        self.add_re_issue_focus == False

    def issue_close(self):
        self.close_reset()
        self.close()
    
    def closeEvent(self,event):
        self.close_reset()
        event.accept()

    def contribution_issue_s(self):  # 기부금영수증 뱔행
        from basic.sungdo import s_name_select
        from register.register_sql import contribution_reg
        global issuable_amount, real_name,hap_code, i_amount, sv_date,ev_date, issu_date
        # conn = pymysql.connect(host = 'localhost', user='root', password='0000', db = 'isbs2024',charset ='utf8')
        # cur = conn.cursor()
        tryno = 0
        is_date = self.issue_date_widget.text()
        issu_date = datetime.strptime(is_date,'%Y-%m-%d')
        hap_code = self.issue_name_hapcode_widget.text()
        name_diff = self.issue_name_widget.text()
        if hap_code != "":
            codename = s_name_select(name_diff)
            r_name = codename[0][1]
            
            tar_year_c = self.target_year_widget.text()
            if (name_diff != "" and tar_year_c != "") or (name_diff == None and tar_year_c == None): 
                # try:
                # if hap_code != '' :
                iss_amount = self.issue_amount_widget.text()
                isss_amount = iss_amount.replace(',','')
                if isss_amount != '' :
                    is_amount = int(isss_amount)
                    if issuable_amount > 0 and  issuable_amount >= is_amount:
                        i_amount = i_amount.replace(',','')
                        jumin = self.jumin_no_widget.text()
                        if jumin != '':
                            jumin = self.jumin_no_widget.text()
                        else:
                            jumin = ''
                        addr = self.address_widget.text()
                        if addr != '':
                            addr = self.address_widget.text()
                        else:
                            addr = ''
                        category = self.category_widget.text()
                        real_name = str(r_name)
                        r_name = real_name.strip("[',']")
                        juj = str(jujung_ratio*100) + "%"

                        cat_code = '41'
                        addr2 = church_addr #'경기도 고양시 일산동구 강송로 125번길 7-2'
                        buss_no = church_biz_no # '606-82-78210'

                        data = (issu_date, issue_No, tar_year, hap_code, name_diff,jumin,addr,category,cat_code,"금전기부",sv_date,ev_date,juj,is_amount,'최초발행',church_name,buss_no,addr2)

                        imsi_data.append([issu_date, issue_No,tar_year,hap_code,name_diff,jumin,addr,category,cat_code,
                                            "금전기부",sv_date,ev_date,juj,is_amount,'최초발행',church_name,buss_no,'',addr2])

                        try:
                            if len(imsi_data) > 0:
                                gibu = load_workbook(r"./print/contribution_recept.xlsx")
                                issue_date = datetime.strftime(issu_date,'%Y-%m-%d')
                                ws = gibu["tax_form"]
                                ws["B3"] = issue_No ; ws["C7"] = real_name ; ws["H7"] = jumin; ws["C8"] = addr 
                                ws["B11"] = church_name ; ws["H11"] = buss_no; ws["B13"] = addr2
                                ws["A22"] = category ; ws["B22"] = cat_code ; ws["D22"] = "금전기부"; ws["E22"] = sv_date; ws["F22"] = ev_date 
                                ws["J22"] = is_amount; ws["J31"] = issue_date; ws["J37"] = issue_date ; ws["I33"] = r_name
                                ws["F40"] = church_name
                                img = Image("./img/img_1.png")
                                img.height = 85 #22.6
                                img.width = 90 #23.9
                                ws.add_image(img, "K38")

                                img_2 = Image("./img/biz_no.jpg")
                                img_2.height = 1000
                                img_2.width = 720 
                                ws.add_image(img_2, "A48")

                                gibu.save(os.path.join(Location, targ_year + "기부금영수증" + '.xlsx'))   # + real_name 를 제거하여 이름없이 저장함
                                gibu.close()

                                self.issue_name_widget.clear()
                                self.issue_name_hapcode_widget.clear()
                                self.target_year_widget.clear()
                                self.issue_amount_widget.clear()
                                self.total_amount_widget.clear()
                                self.issable_amount.clear()
                                # self.issue_church_widget.clear()
                                self.issue_no_widget.clear()
                                self.jumin_no_widget.clear()
                                self.address_widget.clear()
                        
                            self.issue_name_widget.clear()
                            self.issue_amount_widget.clear()
                            # self.issue_church_widget.clear()
                            self.jumin_no_widget.clear()
                            self.address_widget.clear()
                            gibu.close()

                            imsi_data.clear()
                            j = 1
                            # QMessageBox.about(self,'저장',"발행 리스트가 저장되었습니다.!!!")

                            filename = os.path.join(Location, targ_year + "기부금영수증" + '.xlsx')  # + real_name  파일저장에서 이름없이 저장 , 마지막 발행한 것만 남기자
                            # os.popen(filename,'r')  더이상 지원하지 않으므로 아래와 같이 수정함.
                            subprocess.run(["start", "excel", filename], shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, close_fds=True)
                            contribution_reg(data)
                            # os.system("start excel "+ filename)
                        except ValueError: #OSError : #(errno() , strerror[filename[, winerror[,filename2]]]):
                            QMessageBox.about(self, '파일열기 에러',"'contribution_recept.xlsx'파일이 열려 있습니다. 파일을 닫고 다시 진행해 주세요. !!!")
                        
                        except PermissionError: # [Errno 13] Permission denied: 'D:\\Data\\baekseok\\cont_issue\\2023기부금영수증.xlsx'
                            QMessageBox.about(self, '파일저장 에러',"'contribution_recept.xlsx'파일이 열려 있습니다. 파일을 닫고 다시 진행해 주세요. !!!")
                    else:
                        QMessageBox.about(self,'',"발행할 수 있는 잔액을 확인하세요!!") 

                else :
                    QMessageBox.about(self, '금액확인'," 금액이 확인되지 않습니다. 확인후 다시 진행해 주세요. !!!")
                    self.issue_name_widget.clear()
                    self.issue_name_hapcode_widget.clear()
                    self.issue_amount_widget.clear()
                    self.jumin_no_widget.clear()
                    self.address_widget.clear()

                

            else:
                QMessageBox.about(self,'입력오류 !!!','"검색할 이름" 또는 "발행대상연도" 를 확인하여 주십시오')
                self.issue_name_widget.clear()
                self.issue_name_hapcode_widget.clear()
                self.issue_amount_widget.clear()
                self.jumin_no_widget.clear()
                self.address_widget.clear()
        else :
            QMessageBox.about(self,'입력오류 !!!','"검색할 이름" 을 확인하여 주십시오')