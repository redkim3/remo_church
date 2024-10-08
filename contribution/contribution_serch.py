import os, subprocess
from PyQt5.QtCore import Qt 
from PyQt5 import uic 
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIcon

from datetime import datetime
from openpyxl.drawing.image import Image
cur_fold = os.getcwd()

Location = os.path.join(cur_fold, "cont_issue") # 'D:/Data/baekseok/cont_issue/'

form_class = uic.loadUiType(os.path.join(cur_fold,"ui", "contribution_serch_form.ui"))[0]
imsi_data=[]
class contributionSerch(QDialog, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        
        self.setWindowIcon(QIcon(os.path.join(cur_fold,'img','logo.ico')))
        self.setWindowTitle('기부금영수증 발급 검색')

        serch_Button = QPushButton("검색하기")
        serch_Button.clicked.connect(self.serch_contribution)

        end_close = QPushButton("종료")
        end_close.clicked.connect(self.cancel_close)
        self.name_code_widget.text()
        self.name_code_widget.setFocus()
        self.name_code_widget.editingFinished.connect(self.hap_code_select)
        
    def hap_code_select(self):
        from basic.sungdo import s_name_select             #  hap_code_select 
        try:
            self.view_tableWidget.clearContents()
            self.hap_code_widget.clear()
            self.hap_widget.clear()
            #self.target_year_widget.clearContents()
            n_code = self.name_code_widget.text()
            hapcode = s_name_select(n_code)
            hapcode = hapcode[0][0]
            hap_code = hapcode.strip("[',']")
            self.hap_code_widget.setText(hap_code)
        except:
            QMessageBox.about(self,'입력오류 !!!','등록되지 않은 이름 입니다. 확인하여 주십시오')
            self.name_code_widget.clear()
            self.hap_code_widget.clear()

    def serch_contribution(self):
        from basic.contribution_issue import issued_status_serch
        global vdate
        hap_total = 0
        self.view_tableWidget.clearContents()
        hap_code = self.hap_code_widget.text()

        if hap_code != '':
            
            try:
                Y1 = int(self.target_year_widget.text()) # target_year_widget
                issue_detail = issued_status_serch(Y1,hap_code)
                set_row = len(issue_detail)
                self.view_tableWidget.setRowCount(set_row)
                for j in range(set_row):  # j는 행 c는 열
                    dat = issue_detail[j][0]
                    vdate = dat.strftime('%Y-%m-%d')
                    sign = str(issue_detail[j][1])
                    s_name = str(issue_detail[j][2])

                    amo_int = int(issue_detail[j][3])
                    amo = format(amo_int,",")
                    iss_detail = str(issue_detail[j][4])
                    if iss_detail == '최초발행':
                        hap_total += amo_int
                    self.view_tableWidget.setItem(j,0,QTableWidgetItem(vdate))
                    self.view_tableWidget.setItem(j,1,QTableWidgetItem(sign))
                    self.view_tableWidget.setItem(j,2,QTableWidgetItem(s_name))
                    self.view_tableWidget.setItem(j,3,QTableWidgetItem(amo))
                    self.view_tableWidget.setItem(j,4,QTableWidgetItem(iss_detail))
                    self.view_tableWidget.resizeColumnsToContents()
                    self.view_tableWidget.item(j,0).setTextAlignment(Qt.AlignVCenter|Qt.AlignHCenter)
                    self.view_tableWidget.item(j,1).setTextAlignment(Qt.AlignVCenter|Qt.AlignHCenter)
                    self.view_tableWidget.item(j,2).setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                    self.view_tableWidget.item(j,3).setTextAlignment(Qt.AlignVCenter|Qt.AlignRight)
                    self.view_tableWidget.item(j,4).setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)

                hap_total = format(hap_total,",")
                self.hap_widget.setText(hap_total)
                        
            except ValueError:
                QMessageBox.about(self,'검색내용에러',"발급대상년도 가 없습니다.!!!")   
                self.hap_widget.clear()
                self.name_code_widget.clear()
                self.hap_code_widget.clear()
        else:
            QMessageBox.about(self,'검색내용에러',"발급대상 이름이 없습니다.!!!")   
            self.hap_widget.clear()
            self.name_code_widget.clear()
            self.hap_code_widget.clear()

    def cancel_close(self):
        # self.view_tableWidget.clearContents()
        self.view_tableWidget.setRowCount(0) # clear()
        self.name_code_widget.clear()
        self.hap_code_widget.clear()
        self.hap_widget.clear()
        self.target_year_widget.clear()
        self.close()

    def re_issue_contribution(self):
        from basic.contribution_issue import re_issued_list_serch
        from register.register_sql import contribution_re_reg
        
        global vdate,hap_code
        Y1 = self.target_year_widget.text()
        serch_sign = self.serch_sign_widget.text()
        hap_code = self.hap_code_widget.text()
        if serch_sign != '':
            reissue = re_issued_list_serch(serch_sign)

            try:
                i_date =reissue[0][0]  # 발급일
                is_date=datetime.strftime(i_date,'%Y-%m-%d')
                
                sign1=str(reissue[0][1]) # 발급기호
                re_name=str(reissue[0][2]);  ju_no = str(reissue[0][3])  # 신청자이름][ 주민번호
                re_addr= str(reissue[0][4]);  re_cate =str(reissue[0][5]) # 신청자 주소, 유형
                re_cate_code =str(reissue[0][6]);    re_gubun = str(reissue[0][7]) #유형코드, 구분
                s1_date = reissue[0][8];   re_s_date = datetime.strftime(s1_date,'%Y-%m-%d') #시작일
                e1_date = reissue[0][9];   re_e_date = datetime.strftime(e1_date,'%Y-%m-%d')  #종료일
                juj = reissue[0][10];  # 주정비율
                re_amo = int(reissue[0][11]); re_amount = format(re_amo,",") #확인금액, 확인금액(,)
                re_ch_name = str(reissue[0][12]); re_ch_bus_no = str(reissue[0][13]) # 교회이름, 사업자번호
                re_ch_addr = str(reissue[0][14]) # 교회주소
                if ju_no == 'nan':
                    ju_no = None
                if re_addr == "nan":
                    re_addr = None

                imsi_data.append((i_date, sign1, Y1, hap_code, re_name, ju_no, re_addr, re_cate, re_cate_code, re_gubun, s1_date, e1_date, juj, re_amo,'재발행', re_ch_name, re_ch_bus_no, re_ch_addr))
                sending_imsi_data = tuple(imsi_data)
                
                if len(imsi_data) > 0:
                    from openpyxl import load_workbook
                    gibu = load_workbook(r"./print/contribution_recept.xlsx")
                    ws = gibu["tax_form"]

                    ws["B3"] = sign1; ws["C7"] = re_name
                    ws["H7"] = ju_no;  ws["C8"] = re_addr
                    ws["B11"] = re_ch_name; ws["H11"] = re_ch_bus_no; ws["B13"] = re_ch_addr
                    ws["A22"] = re_cate; ws["B22"] = re_cate_code; ws["D22"] = re_gubun; 
                    ws["E22"] = re_s_date; ws["F22"] = re_e_date
                    ws["J22"] = re_amount; ws["I33"] = re_name
                    ws["J31"] = is_date; ws["J37"] = is_date ; 
                    ws["F40"] = re_ch_name

                    img = Image("./img/img_1.png")
                    img.height = 85 #22.6
                    img.width = 90 #23.9
                    ws.add_image(img, "K38")

                    img_2 = Image("./img/biz_no.jpg")
                    img_2.height = 1000
                    img_2.width = 720 
                    ws.add_image(img_2, "A48")

                    imsi_data.clear()
                    
                    gibu.save(os.path.join(Location, Y1 + "기부금영수증_재" + '.xlsx'))
                    gibu.close()

                    filename = os.path.join(Location, Y1 + "기부금영수증_재" + '.xlsx')
                    subprocess.run(["start", "excel", filename], shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, close_fds=True)
                    contribution_re_reg(sending_imsi_data)
                    self.serch_sign_widget.clear()
                
            except OSError : #(errno() , strerror[filename[, winerror[,filename2]]]):
                QMessageBox.about(self,'파일열기 에러',"'파일이 열려 있습니다. 파일을 닫고 다시 진행해 주세요. !!!")    
                imsi_data.clear()
        else:
            QMessageBox.about(self,'없음',"'재발행할 발급기호를 넣어 주세요. !!!")    
