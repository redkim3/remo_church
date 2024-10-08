import os, subprocess, datetime # sys, psutil
from PyQt5.QtCore import *
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import *
from openpyxl.styles import Font
from PyQt5 import uic

today = QDate.currentDate()
cur_fold = os.getcwd()
form_class = uic.loadUiType("./ui/week_ge_bogo.ui")[0]

class weekly_Ge_report(QDialog, form_class) : #QDialog
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        global ws, bogoseo
        self.setWindowTitle("일반회계 재정보고")
        self.setWindowIcon(QIcon(os.path.join(cur_fold,'img','logo.ico')))
        bogo_year = today.year()
        bogo_month = today.month()
        self.bogo_year_widget.setText(str(bogo_year))
        self.bogo_month_widget.setText(str(bogo_month))
        self.bogo_week_widget.setText("1")
        self.bogo_month_widget.setFocus()
    
    def button(self):
        self.create_button.clicked.disconnect(self.Y1_M1_W1_execute)
        self.bogo_year_widget.editingFinished.connect(self.confirmclear)
        self.bogo_month_widget.editingFinished.connect(self.confirmclear)
        self.bogo_week_widget.editingFinished.connect(self.confirmclear)

    def confirmclear(self):
        self.week_bogo_confirm.clear()

    def Y1_M1_W1_execute(self):
        from basic.hun_split import today_income_report, today_hun_list, bank_income_call, past_income_sum 
        from basic.cost_split import past_cost_by_year_week, today_cost_list
        from openpyxl import load_workbook
        global Y1, M1, W1
        try:
            bogoseo = load_workbook(r"./print/bogoseo.xlsx")
            hun_income_mok = []; today_income_sum = 0; bank_income_sum = 0; ge_bank_income_sum = 0; j = 0; ban = 0
            Y1 = int(self.bogo_year_widget.text())
            M1 = int(self.bogo_month_widget.text())
            W1 = int(self.bogo_week_widget.text())

            ws = bogoseo["회계재정보고"]
            for row_1 in ws['A8:I40']:
                for cell in row_1:
                    cell.value = None
            for row_2 in ws['B51:I89']:
                for cell in row_2:
                    cell.value = None
            for row_3 in ws['N51:O90']:
                for cell in row_3:
                    cell.value = None
            for row_4 in ws['Q51:R90']:
                for cell in row_4:
                    cell.value = None
            for row_5 in ws['T51:U90']:
                for cell in row_5:
                    cell.value = None
            ws["H90"] = None; ws["O49"] = None; ws['N46'] = None
            ws["D4"] = None

            ws["A2"] = Y1
            ws["A2"].font = Font(size=18, bold=True)
            ws["B3"] = "일반회계 주간 재정보고"
            ws["B3"].font = Font(size=22, bold=True)
            ws["C5"] = "{0}월 {1}주".format(M1,W1) 
            ws["C5"].font = Font(size=12, bold=True)
        
            today_income = today_income_report(Y1,M1,W1)
            
            # 결과 출력
            for row in today_income:  # 총 입금액
                mok = str(row[0]); amo = int(row[1])
                ws["A" + str(8 + j)] = mok
                hun_income_mok.append(mok)  # 여기서 append 사용
                ws["C" + str(8 + j)] = amo
                today_income_sum += amo  # 금일 수입 총액 
                j += 1

            Bank_income = bank_income_call(Y1,M1,W1)
                
            for row in Bank_income:  # 통장예입
                hun_mok, amount = row
                if amount != 0:
                    ban += 1
                    ws[("N"+ str(9+ban))] = hun_mok
                    ws[("O"+ str(9+ban))] = amount
                    if hun_mok != '선교헌금':
                        ge_bank_income_sum = ge_bank_income_sum + row[1]    
                    bank_income_sum = bank_income_sum + row[1]   #금일 통장입금액

            
            hun_list_source = today_hun_list(Y1, M1, W1)

            # 수정 시작
            hun_mok_list = [row[0] for row in hun_list_source]
            name_diff_list = [row[2].split(',') if row[2] else [] for row in hun_list_source]
            hun_hang_list = [row[1] for row in hun_list_source]
            amount_list = [row[3] for row in hun_list_source]
            hun_detail_list = [row[4] for row in hun_list_source]
            bank_list = [row[5] for row in hun_list_source]
            marks_list = [row[6] if len(row) > 6 else None for row in hun_list_source]
            # 리스트들을 zip으로 묶어서 정렬
            sorted_data = sorted(zip(hun_mok_list,name_diff_list, hun_hang_list, amount_list, hun_detail_list, bank_list, marks_list), key=lambda x: float(x[3]), reverse=True)
            hun_source = [(item[0], item[1][0] if item[1] and isinstance(item[1], list) else item[1], item[2], item[3] if item[3] != None else None, item[4], item[5], item[6]) for item in sorted_data]
            hun_list = []
            for item in hun_source:
                hun_list.append(item)
            
            sip_list = []; gam_list = []; jujeong_list = []; jeolgi_list = []; other_list = []; mission_data = []; jijeong_list = []; mokjeok_list = []
            mission_list = []  # mission_list는 mission_data를 정렬한 리스트임
            for hun_row in hun_list:
                if hun_row[0] == '십일조헌금':
                    sip_list.append((hun_row[0],hun_row[1],hun_row[2],hun_row[3],hun_row[4],hun_row[5]))
                elif hun_row[0] == '감사헌금':
                    gam_list.append((hun_row[0],hun_row[1],hun_row[2],hun_row[3],hun_row[4],hun_row[5]))
                elif hun_row[0] == '주일헌금':
                    jujeong_list.append((hun_row[0],hun_row[1],hun_row[2],hun_row[3],hun_row[4],hun_row[5]))
                elif hun_row[2] == '절기헌금':
                    jeolgi_list.append((hun_row[0],hun_row[1],hun_row[2],hun_row[3],hun_row[4],hun_row[5]))
                elif hun_row[2] == '선교헌금':
                    mission_data.append((hun_row[0],hun_row[1],hun_row[2],hun_row[3],hun_row[4],hun_row[5]))
                elif hun_row[2] == '지정헌금' and hun_row[0] == '목적헌금':
                    mokjeok_list.append((hun_row[0],hun_row[1],hun_row[2],hun_row[3],hun_row[4],hun_row[5]))
                elif hun_row[2] == '지정헌금' and hun_row[0] != '목적헌금':
                    jijeong_list.append((hun_row[0],hun_row[1],hun_row[2],hun_row[3],hun_row[4],hun_row[5]))
                else:                      # 기타소득
                    if hun_row[2] == '기타소득':
                        other_list.append((hun_row[0],hun_row[1],hun_row[2],hun_row[3],hun_row[4],hun_row[5],hun_row[6]))
            
            custom_order = ['선교헌금', '구역헌금', '이자소득_선교']
            
            def custom_sort(item):
                order = {value: index for index, value in enumerate(custom_order)}
                return order.get(item[0], len(custom_order))
            
            # 사용자 정의 함수를 key로 사용하여 정렬
            mission_list = sorted(mission_data, key=custom_sort)

            sip_sum1 = 0; sip_sum2 = 0
            a = 0; aa = 0; b = 0; bb = 0; c = 0; cc = 0;ccc = 0; d = 0; dd = 0; e = 0; ee = 0
            jc = 0; jc1 = 0; jc2 = 0; fx = 0; jj1 = 0; jj2 = 0; jj3 = 0; jj4 = 0
            for sip in sip_list: # 십일조헌금 엑셀 출력
                if a < 39 :
                    ws[("B"+ str(49))] = sip[0]  # 헌금명칭
                    ws[("B"+ str(90))] = '소 계'
                    if sip[5] == '통장예입':
                        ws[("B"+ str(51+a))] = "*" + str(sip[1])
                        ws[("C"+ str(51+a))] = sip[3]
                    else:
                        ws[("B"+ str(51+a))] = str(sip[1])
                        ws[("C"+ str(51+a))] = sip[3]  
                    sip_sum1 += sip[3]
                    ws[("B"+ str(90))] = '소 계'
                    ws[("C"+ str(90))] = sip_sum1
                    a += 1 
                else:
                    ws[("A"+ str(92))] = '헌금자 명단 보고서(2)'
                    ws[("B"+ str(96))] = sip[0]
                    ws[("B"+ str(97))] = '성 명'; ws[("C"+ str(97))] = '금 액'
                    if sip[5] == '통장예입':
                        ws[("B"+ str(98+aa))] = "*" + sip[1]
                        ws[("C"+ str(98+aa))] = sip[3]
                    else:
                        ws[("B"+ str(98+aa))] = sip[1]
                        ws[("C"+ str(98+aa))] = sip[3]
                    sip_sum2 += sip[3]
                    ws[("B"+ str(137))] = '소 계'
                    ws[("C"+ str(137))] = sip_sum1
                    aa += 1
                    a += 1      # 여기까지 십일조 출력
                    
            gam_sum1 = 0; gam_sum2 = 0
            for gam in gam_list: # 감사헌금 엑셀출력
                if b < 39:
                    ws[("D"+ str(49))] = gam[0]  # 금액은 E
                    ws[("D"+ str(90))] = '소 계'
                    if gam[5] == '통장예입':
                        ws[("D"+ str(51+b))] = "*" + str(gam[1])
                        ws[("E"+ str(51+b))] = gam[3]
                    else:
                        ws[("D"+ str(51+b))] = str(gam[1])
                        ws[("E"+ str(51+b))] = gam[3]
                    gam_sum1 += gam[3]  
                    ws[("D"+ str(90))] = '소 계'
                    ws[("E"+ str(90))] = gam_sum1
                    b += 1
                else:
                    ws[("A"+ str(92))] = '헌금자 명단 보고서(2)'
                    ws[("D"+ str(96))] = 'row[0]'
                    ws[("D"+ str(97))] = '성 명'; ws[("E"+ str(97))] = '금 액'
                    if gam[5] == '통장예입':
                        ws[("D"+ str(98+bb))] = "*" + str(gam[1])
                        ws[("E"+ str(98+bb))] = gam[3]
                    else:
                        ws[("D"+ str(98+bb))] = str(gam[1])
                        ws[("E"+ str(98+bb))] = gam[3]
                        gam_sum2 += gam[3]
                        ws[("D"+ str(137))] = '소 계'
                        ws[("E"+ str(137))] = gam_sum2
                    bb += 1
                    b += 1
            jeol_sum1 = 0; jeol_sum2 = 0; jeol_sum3 = 0
            for jeol in jeolgi_list:   # 절기 헌금
                if len(jeolgi_list) <= 7:
                    ws[("A"+ str(33))] = jeol[0] # 금액은 H
                    if jeol[5] == '통장예입':
                        ws[("B"+ str(34+ccc))] = "*" + str(jeol[1])
                        ws[("C"+ str(34+ccc))] = jeol[3]
                    else:
                        ws[("B"+ str(34+ccc))] = str(jeol[1])
                        ws[("C"+ str(34+ccc))] = jeol[3]
                    jeol_sum1 += jeol[3]
                    ws[("C"+ str(33))] = jeol_sum1
                    ccc += 1
                else:
                    if 7 < len(jeolgi_list) < 39:
                        ws[("F"+ str(49))] = jeol[0] # 금액은 H
                        ws[("F"+ str(90))] = '소 계'
                        if jeol[5] == '통장예입':
                            ws[("F"+ str(51+c))] = "*" + str(jeol[1])
                            ws[("H"+ str(51+c))] = jeol[3]
                        else:
                            ws[("F"+ str(51+c))] = str(jeol[1])
                            ws[("H"+ str(51+c))] = jeol[3]
                        jeol_sum1 += jeol[3]
                        ws[("F"+ str(90))] = '소 계'
                        ws[("H"+ str(90))] = jeol_sum1
                        c += 1
                    elif len(jeolgi_list) >= 39:

                        ws[("N"+ str(46))] = '{} 명단 보고서'.format(jeol[0])
                        ws[("N"+ str(50))] = '성 명'; ws[("O"+ str(50))] = '금 액'
                        ws[("Q"+ str(50))] = '성 명'; ws[("R"+ str(50))] = '금 액'
                        ws[("T"+ str(50))] = '성 명'; ws[("V"+ str(50))] = '금 액'

                        if jc < 40:
                            if jeol[5] == '통장예입':
                                ws[("N"+ str(51+jc))] = "*" + str(jeol[1])
                                ws[("O"+ str(51+jc))] = jeol[3]
                            else:
                                ws[("N"+ str(51+jc))] = str(jeol[1])
                                ws[("O"+ str(51+jc))] = jeol[3]
                            jeol_sum1 += jeol[3]
                            ws[("N"+ str(90))] = '소 계'
                            ws[("O"+ str(90))] = jeol_sum1
                            jc += 1
                        elif  39 < jc < 79 :
                            if jeol[5] == '통장예입':
                                ws[("Q"+ str(51 + jc1))] = "*" + str(jeol[1])
                                ws[("R"+ str(51 + jc1))] = jeol[3]
                            else:
                                ws[("Q"+ str(51 + jc1))] = str(jeol[1])
                                ws[("R"+ str(51 + jc1))] = jeol[3]
                            jeol_sum2 += jeol[3]
                            ws[("Q"+ str(90))] = '소 계'
                            ws[("R"+ str(90))] = jeol_sum2
                            jc += 1
                            jc1 += 1
                        else:
                            if jeol[5] == '통장예입':
                                ws[("T"+ str(51 + jc2))] = "*" + str(jeol[1])
                                ws[("V"+ str(51 + jc2))] = jeol[3]
                            else:
                                ws[("T"+ str(51 + jc2))] = str(jeol[1])
                                ws[("V"+ str(51 + jc2))] = jeol[3]
                            jeol_sum3 += jeol[3]
                            ws[("T"+ str(90))] = '소 계'
                            ws[("V"+ str(90))] = jeol_sum3
                            jc += 1
                            jc2 += 1
                    ws[("O"+ str(49))] = jeol_sum1 + jeol_sum2 + jeol_sum3

            seon_sum1 = 0 
            for seon in mission_list:
                if c == 0 and len(mission_list) < 39:
                    ws[("F"+ str(49))] = seon[2]  # 금액은 H
                    ws[("F"+ str(90))] = '소 계'
                    if seon[5] == '통장예입':
                        ws[("F"+ str(51+d))] = "*" + str(seon[1])
                        ws[("H"+ str(51+d))] = seon[3]
                    else:
                        ws[("F"+ str(51+d))] = str(seon[1])
                        ws[("H"+ str(51+d))] = seon[3]
                    seon_sum1 += seon[3]
                    ws[("F"+ str(90))] = '소 계'
                    ws[("H"+ str(90))] = seon_sum1
                    d += 1

            other_sum = 0
            for other in other_list:
                if len(other_list) > 0 :
                    ws[("A"+ str(17))] = other[2]  # 금액은 H
                    if other[5] == '통장예입':
                        ws[("B"+ str(18+e))] = str(other[6])
                        ws[("C"+ str(18+e))] = other[3]
                    else:
                        ws[("B"+ str(18+e))] = str(other[6])
                        ws[("C"+ str(18+e))] = other[3]
                    other_sum += other[3]
                    e += 1
                    ws[("C"+ str(17))] = other_sum
            mokjeok_sum1 = 0 ; mo1 = 0
            for mokjeok in mokjeok_list:   # 지정 헌금
                if len(mokjeok_list) + len(other_list) <= 10:
                    mo1 += 1
                    if len(other_list) == 0:
                        mok_cn = 0
                    else:
                        mok_cn = len(other_list) + 1
                    ws[("A"+ str(17 + mok_cn))] = mokjeok[0] # hun_detail
                    if mokjeok[5] == '통장예입':
                        ws[("B"+ str(17+mok_cn+mo1))] = "*" + str(mokjeok[1])+str(mokjeok[4])
                        ws[("C"+ str(19+mok_cn+mo1))] = mokjeok[3]
                    else:
                        ws[("B"+ str(17+mok_cn+mo1))] = str(mokjeok[1])+str(mokjeok[4])
                        ws[("C"+ str(17+mok_cn+mo1))] = mokjeok[3]
                    mokjeok_sum1 += mokjeok[3]
                    ws[("C"+ str(17+mok_cn))] = mokjeok_sum1

            jijeong_sum1 = 0 ; jijeong_sum2 = 0; jijeong_sum3 = 0; fn = 1
            for jijeong in jijeong_list:   # 지정 헌금
                if len(jijeong_list) + len(mokjeok_list) + len(other_list) <= 10:
                    jj1 += 1
                    if len(other_list) == 0:
                        if len(mokjeok_list) == 0 :
                            fn = 0
                        else:
                            fn = len(other_list) + 1
                    else:
                        if len(mokjeok_list) == 0 :
                            fn = len(other_list) + 1
                        else:
                            fn = len(other_list) + len(mokjeok_list) + 2

                    ws[("A"+ str(17 + fn))] = jijeong[0] # 금액은 H
                    if jijeong[5] == '통장예입':
                        ws[("B"+ str(17+fn+jj1))] = "*" + str(jijeong[1])
                        ws[("C"+ str(17+fn+jj1))] = jijeong[3]
                    else:
                        ws[("B"+ str(17+fn+jj1))] = str(jijeong[1])
                        ws[("C"+ str(17+fn+jj1))] = jijeong[3]
                    jijeong_sum1 += jijeong[3]
                    ws[("C"+ str(17+fn))] = jijeong_sum1
                
                elif len(jijeong_list) >= 10:
                    ws[("N"+ str(46))] = '{} 명단 보고서'.format(jijeong[0])
                    ws[("N"+ str(50))] = '성 명'; ws[("O"+ str(50))] = '금 액'
                    ws[("Q"+ str(50))] = '성 명'; ws[("R"+ str(50))] = '금 액'
                    ws[("T"+ str(50))] = '성 명'; ws[("V"+ str(50))] = '금 액'
                    if jj2 < 40:
                        if jijeong[5] == '통장예입':
                            ws[("N"+ str(51+jj2))] = "*" + str(jijeong[1])
                            ws[("O"+ str(51+jj2))] = jijeong[3]
                        else:
                            ws[("N"+ str(51+jj2))] = str(jijeong[1])
                            ws[("O"+ str(51+jj2))] = jijeong[3]
                        jijeong_sum1 += jijeong[3]
                        ws[("N"+ str(90))] = '소 계'
                        ws[("O"+ str(90))] = jijeong_sum1
                        jj2 += 1
                    elif  39 < jj2 < 79 :
                        if jijeong[5] == '통장예입':
                            ws[("Q"+ str(51+jj3))] = "*" + str(jijeong[1])
                            ws[("R"+ str(51+jj3))] = jijeong[3]
                        else:
                            ws[("Q"+ str(51+jj3))] = str(jijeong[1])
                            ws[("R"+ str(51+jj3))] = jijeong[3]
                        jijeong_sum2 += jijeong[3]
                        ws[("Q"+ str(90))] = '소 계'
                        ws[("R"+ str(90))] = jijeong_sum2
                        jc += 1
                        jj3 += 1
                    elif  79 <= jj2 < 118 :
                        if jijeong[5] == '통장예입':
                            ws[("T"+ str(51+jj4))] = "*" + str(jijeong[1])
                            ws[("V"+ str(51+jj4))] = jijeong[3]
                        else:
                            ws[("T"+ str(51+jj4))] = str(jijeong[1])
                            ws[("V"+ str(51+jj4))] = jijeong[3]
                        jijeong_sum3 += jijeong[3]
                        ws[("T"+ str(90))] = '소 계'
                        ws[("V"+ str(90))] = jijeong_sum3
                        jj2 += 1
                        jj4 += 1
                    ws[("O"+ str(49))] = jijeong_sum1 + jijeong_sum2 + jijeong_sum3
 
            #  지출내역 인쇄
            non_in_bank = 0; cash = 0 ; today_cost_value = 0
            today_cost = today_cost_list(Y1,M1,W1)
            past_cost = past_cost_by_year_week(Y1,M1,W1)
            cos1 = 0
            for cost_row in today_cost:
                ws[("D"+ str(8+cos1))] = str(cost_row[0])
                ws[("E"+ str(8+cos1))] = cost_row[1]
                if cost_row[3] == '자동이체' or cost_row[3] == '현금':
                    non_in_bank += cost_row[1]
                if cost_row[3] == '현금':
                    cash += cost_row[1]
                today_cost_value += cost_row[1]
                if cost_row[2] != None :
                    ws[("F"+ str(8+cos1))] = str(cost_row[2])
                cos1 += 1
            ws["A41"]='일반회계 통장예입'
            ws["C" + str(41)] = ge_bank_income_sum
            ws["D41"]='금주수입총액'
            ws["E" + str(41)] = today_income_sum
            ws["A42"]='일반회계 현금수입'
            ws["C" + str(42)] = today_income_sum - ge_bank_income_sum
            past_income_amount = past_income_sum(Y1,M1,W1)  # 과거 누적 수입금액- 당일 금액 제외
            ws["A43"] = "지난주 잔액"
            ws["C" + str(43)] = past_income_amount - past_cost
            ws["D42"] = "금주지출총액"
            ws["E" + str(42)] = today_cost_value
            ws["D43"] = "금주 차액"
            ws["E" + str(43)] = today_income_sum - today_cost_value
            ws["F41"] = "현금 및 자동이체"
            ws["H" + str(41)] = non_in_bank
            ws["F42"] = "입 금 예 정 액"
            ws["H" + str(42)] = today_income_sum - ge_bank_income_sum - cash
            ws["F43"] = "금주 마감잔액"
            ws["H" + str(43)] = (past_income_amount+today_income_sum) - (past_cost+today_cost_value)

            bogoseo.save('./excel_view/일반회계_재정보고.xlsx')
            self.week_bogo_confirm.setText("{}년도 {}월 {}주 일반회계 재정보고서가 생성되었습니다.".format(Y1,M1,W1))
            bogoseo.close()
            
            subprocess.Popen(["start", "excel.exe", os.path.abspath("./excel_view/일반회계_재정보고.xlsx")], shell=True)
            
        except PermissionError : #(errno() , strerror[filename[, winerror[,filename2]]]): #OSError
                QMessageBox.about(self,'파일열기 에러',"'일반회계_재정보고.xlsx'파일이 열려 있습니다. 파일을 닫고 다시 진행해 주세요. !!!")
        except ValueError: #invalid literal for int() with base 10
            QMessageBox.about(self,'검색 에러',"'검색대상 년도, 월, 주 에대한 입력을 확인해 주세요. !!!")
