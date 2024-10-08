import os, subprocess
from PyQt5.QtCore import QDate
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import *
from PyQt5 import uic

today = QDate.currentDate()
cur_fold = os.getcwd()
form_class = uic.loadUiType("./ui/week_sun_bogo.ui")[0]

class weekly_mission_report(QDialog, form_class) : #QDialog
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("선교회계 재정보고")
        self.setWindowIcon(QIcon(os.path.join(cur_fold,'img','logo.ico')))
        global ws, bogoseo, Y1, M1, W1
        bogo_year = str(today.year())
        bogo_month = str(today.month())
        self.bogo_year_widget.setText(bogo_year)
        self.bogo_month_widget.setText(bogo_month)
        self.bogo_week_widget.setText("1")
        self.bogo_month_widget.setFocus()

    def button(self):
        self.creat_button.clicked.connect(self.Y1_M1_W1_excute)
        
    def Y1_M1_W1_excute(self):
        from openpyxl import load_workbook
        from basic.hun_split import past_sun_income, today_sun_income_list 
        from basic.cost_split import past_mission_cost, today_mission_cost_list
        bogoseo = load_workbook(r"./print/sun_bogoseo.xlsx")
        ws = bogoseo["선교회계재정보고"]
        for row_1 in ws['A7:H40']:
            for cell in row_1:
                cell.value = None
        ws["B4"] = None

        try:
            for row_1 in ws['A7:H79']:
                for cell in row_1:
                    cell.value = None
            ws["B4"] = None
            self.week_bogo_confirm.clear()
            Y1 = int(self.bogo_year_widget.text())
            M1 = int(self.bogo_month_widget.text())
            W1 = int(self.bogo_week_widget.text())

            past_sun_income_amount = past_sun_income(Y1,M1,W1)
            past_sun_cost_amount = past_mission_cost(Y1,M1,W1)  #:sun_offer_past(Y1, M1, W1)  
            # today_sun_cost_detail = today_mission_cost_list(Y1,M1,W1) #today_sun_offer_detail(Y1, M1, W1)
            
            period = "{0}월 {1}주".format(M1,W1) 
            ws["A2"] = Y1
            ws["B4"] = period
            ws["A5"] = '선교헌금' ; ws["C5"] = '구역헌금'; ws["E5"] = '지출내역'
            ws["A6"] = '성  명' ; ws["B6"] = '금  액'; ws["C6"] = '성  액'
            ws["D6"] = '금  액' ; ws["E6"] = '내  역'; ws["G6"] = '금  액'

            a = 0; a2 = 0; a3 = 0; a4 = 0; b = 0; c = 0; a1 = 0
            b1 = 0; b2 = 0; b3 = 0
            cash_income = 0; Bank_income = 0; today_sun_cost_sum = 0; sun_sum = 0
            seon_sum1 = 0; seon_sum2 = 0; seon_sum3 = 0; seon_sum4 = 0; area_sum = 0

            sun_list_source = today_sun_income_list(Y1,M1,W1)


            # 수정 시작
            sun_mok_list = [row[0] for row in sun_list_source]
            name_diff_list = [row[2].split(',') if row[2] else [] for row in sun_list_source]
            hun_hang_list = [row[1] for row in sun_list_source]
            amount_list = [row[3] for row in sun_list_source]
            hun_detail_list = [row[4] for row in sun_list_source]
            bank_list = [row[5] if len(row) > 5 else None for row in sun_list_source]

            # 리스트들을 zip으로 묶어서 정렬
            sorted_data = sorted(zip(sun_mok_list,name_diff_list, hun_hang_list, amount_list, hun_detail_list, bank_list), key=lambda x: float(x[3]), reverse=True)
            sun_source = [(item[0], item[1][0] if item[1] and isinstance(item[1], list) else item[1], item[2], item[3] if item[3] != None else None, item[4], item[5]) for item in sorted_data]
            sun_list = []

            for item in sun_source:
                sun_list.append(item)

            seonkyo_list = []; area_list = []; other_list = []

            for sun_row in sun_list:
                if sun_row[0] == '선교헌금':
                    seonkyo_list.append((sun_row[0],sun_row[1],sun_row[2],sun_row[3],sun_row[4],sun_row[5]))
                elif sun_row[0] == '구역헌금':
                    area_list.append((sun_row[0],sun_row[1],sun_row[2],sun_row[3],sun_row[4],sun_row[5]))
                else:
                    if sun_row[0] != '선교헌금' and sun_row[0] != '구역헌금':
                        other_list.append((sun_row[0],sun_row[1],sun_row[2],sun_row[3],sun_row[4],sun_row[5]))
            
            ws["A5"].value = '선교헌금'; ws["C5"].value = '구역헌금 외'; ws["E5"].value = '지출내역'

            for seon in seonkyo_list:
                if len(area_list) < 1 :
                    if a < 29 :
                        ws[("A"+ str(5))] = seon[0]  # 헌금명칭
                        ws[("A"+ str(6))] = '성 명'
                        ws[("B"+ str(6))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("A"+ str(7+a))] = "*" + str(seon[1])
                            ws[("B"+ str(7+a))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("A"+ str(7+a))] = str(seon[1])
                            ws[("B"+ str(7+a))] = seon[3]
                            cash_income += seon[3]
                        seon_sum1 += seon[3]
                        ws[("A"+ str(36))] = '소 계'
                        ws[("B"+ str(36))] = seon_sum1
                        a += 1 
                    elif a > 28 and a2 < 29 :
                        ws[("C"+ str(5))] = seon[0]  # 헌금명칭
                        ws[("C"+ str(6))] = '성 명'
                        ws[("D"+ str(6))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("C"+ str(7+a2))] = "*" + str(seon[1])
                            ws[("D"+ str(7+a2))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("C"+ str(7+a2))] = str(seon[1])
                            ws[("D"+ str(7+a2))] = seon[3]
                            cash_income += seon[3]
                        seon_sum2 += seon[3]
                        ws[("C"+ str(36))] = '소 계'
                        ws[("D"+ str(36))] = seon_sum2
                        a += 1 
                        a2 += 1
                    elif a > 56 and a2 > 28 and a3 < 34:
                        ws[("A"+ str(40))] = seon[0]
                        ws[("B"+ str(41))] = '성 명'; ws[("D"+ str(41))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("A"+ str(42+a3))] = "*" + seon[1]
                            ws[("B"+ str(42+a3))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("A"+ str(42+a3))] = seon[1]
                            ws[("B"+ str(42+a3))] = seon[3]
                            cash_income += seon[3]
                        seon_sum3 += seon[3]
                        ws[("A"+ str(78))] = '소 계'
                        ws[("B"+ str(78))] = seon_sum3
                        a += 1
                        a2 += 1
                        a3 += 1   # 여기까지  출력
                    else:
                        if a > 56 and a2 > 28 and a3 >34 and a4 < 35:
                            ws[("C"+ str(40))] = seon[0]
                            ws[("D"+ str(41))] = '성 명'; ws[("D"+ str(41))] = '금 액'
                            if seon[4] == '통장예입':
                                ws[("C"+ str(42+a4))] = "*" + seon[1]
                                ws[("D"+ str(42+a4))] = seon[3]
                                Bank_income += seon[3]
                            else:
                                ws[("C"+ str(42+a4))] = seon[1]
                                ws[("D"+ str(42+a4))] = seon[3]
                                cash_income += seon[3]
                            seon_sum4 += seon[3]
                            ws[("C"+ str(78))] = '소 계'
                            ws[("D"+ str(78))] = seon_sum4
                            a4 += 1
                            a3 += 1
                            a2 += 1
                            a += 1      # 여기까지 십일조 출력

                elif len(area_list) > 15:
                    if a < 28 :
                        ws[("A"+ str(5))] = seon[0]  # 헌금명칭
                        ws[("A"+ str(6))] = '성 명'
                        ws[("B"+ str(6))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("A"+ str(7+a))] = "*" + str(seon[1])
                            ws[("B"+ str(7+a))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("A"+ str(7+a))] = str(seon[1])
                            ws[("B"+ str(7+a))] = seon[3]
                            cash_income += seon[3]
                        seon_sum1 += seon[3]
                        ws[("A"+ str(36))] = '소 계'
                        ws[("B"+ str(36))] = seon_sum1
                        a += 1 
                    elif a2 < 28 :
                        ws[("A"+ str(40))] = seon[0]  # 헌금명칭
                        ws[("A"+ str(41))] = '성 명'
                        ws[("B"+ str(41))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("A"+ str(42+a2))] = "*" + str(seon[1])
                            ws[("B"+ str(42+a2))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("A"+ str(42+a2))] = str(seon[1])
                            ws[("B"+ str(42+a2))] = seon[3]
                            cash_income += seon[3]
                        seon_sum2 += seon[3]
                        ws[("A"+ str(36))] = '소 계'
                        ws[("B"+ str(36))] = seon_sum2
                        a += 1 
                        a2 += 1
                    elif a3 < 34:
                        ws[("C"+ str(40))] = seon[0]
                        ws[("D"+ str(41))] = '성 명'; ws[("D"+ str(41))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("C"+ str(42+a3))] = "*" + seon[1]
                            ws[("D"+ str(42+a3))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("C"+ str(42+a3))] = seon[1]
                            ws[("D"+ str(42+a3))] = seon[3]
                            cash_income += seon[3]
                        seon_sum3 += seon[3]
                        ws[("C"+ str(78))] = '소 계'
                        ws[("D"+ str(78))] = seon_sum3
                        a3 += 1
                        a2 += 1
                        a += 1      # 여기까지 십일조 출력
                elif len(area_list) < 15:
                    if a < 28 :
                        ws[("A"+ str(5))] = seon[0]  # 헌금명칭
                        ws[("A"+ str(6))] = '성 명'
                        ws[("B"+ str(6))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("A"+ str(7+a))] = "*" + str(seon[1])
                            ws[("B"+ str(7+a))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("A"+ str(7+a))] = str(seon[1])
                            ws[("B"+ str(7+a))] = seon[3]
                            cash_income += seon[3]
                        seon_sum1 += seon[3]
                        ws[("A"+ str(36))] = '소 계'
                        ws[("B"+ str(36))] = seon_sum1
                        a += 1 
                    elif a > 28 and (a2 + len(area_list)) < 28 :
                        ws[("C"+ str(len(area_list) + 4))] = seon[0]  # 헌금명칭
                        ws[("C"+ str(len(area_list) + 5))] = '성 명'
                        ws[("D"+ str(len(area_list) + 5))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("C"+ str(len(area_list) + 5 + a2))] = "*" + str(seon[1])
                            ws[("D"+ str(len(area_list) + 5 + a2))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("C"+ str(len(area_list) + 5 + a2))] = str(seon[1])
                            ws[("D"+ str(len(area_list) + 5 + a2))] = seon[3]
                            cash_income += seon[3]
                        seon_sum2 += seon[3]
                        ws[("C"+ str(36))] = '소 계'
                        ws[("D"+ str(36))] = seon_sum2
                        a += 1 
                        a2 += 1
                    elif a > 56 and (len(area_list) + 5 + a2) > 28 and a3 < 34:
                        ws[("A"+ str(40))] = seon[0]
                        ws[("B"+ str(41))] = '성 명'; ws[("D"+ str(41))] = '금 액'
                        if seon[4] == '통장예입':
                            ws[("A"+ str(42+a3))] = "*" + seon[1]
                            ws[("B"+ str(42+a3))] = seon[3]
                            Bank_income += seon[3]
                        else:
                            ws[("A"+ str(42+a3))] = seon[1]
                            ws[("B"+ str(42+a3))] = seon[3]
                            cash_income += seon[3]
                        seon_sum3 += seon[3]
                        ws[("A"+ str(78))] = '소 계'
                        ws[("B"+ str(78))] = seon_sum3
                        a3 += 1
                        a2 += 1
                        a += 1      # 여기까지 십일조 출력
                    else:
                        if a > 56 and (len(area_list) + 5 + a2) > 28 and a3 >34 and a4 < 34:
                            ws[("C"+ str(40))] = seon[0]
                            ws[("D"+ str(41))] = '성 명'; ws[("D"+ str(41))] = '금 액'
                            if seon[4] == '통장예입':
                                ws[("C"+ str(42+a4))] = "*" + seon[1]
                                ws[("D"+ str(42+a4))] = seon[3]
                                Bank_income += seon[3]
                            else:
                                ws[("C"+ str(42+a4))] = seon[1]
                                ws[("D"+ str(42+a4))] = seon[3]
                                cash_income += seon[3]
                            seon_sum4 += seon[3]
                            ws[("C"+ str(78))] = '소 계'
                            ws[("D"+ str(78))] = seon_sum4
                            a4 += 1
                            a3 += 1
                            a2 += 1
                            a += 1      # 여기까지 십일조 출력
            # 구역헌금
            for area in area_list:
                    if b < 29 :
                        ws[("C"+ str(5))] = area[0]  # 헌금명칭
                        ws[("C"+ str(6))] = '성 명'
                        ws[("D"+ str(6))] = '금 액'
                        if area[4] == '통장예입':
                            ws[("C"+ str(7+b))] = "*" + str(area[1])
                            ws[("D"+ str(7+b))] = area[3]
                            Bank_income += area[3]
                        else:
                            ws[("C"+ str(7+b))] = str(area[1])
                            ws[("D"+ str(7+b))] = area[3]
                            cash_income += area[3]
                        area_sum += area[3]
                        ws[("C"+ str(36))] = '소 계'
                        ws[("D"+ str(36))] = area_sum
                        b += 1 
                    
            # 선교헌금 기타
            for other in other_list:
                if len(seonkyo_list) < 30 and len(area_list) < 30 :
                    if other[4] == '통장예입':
                            ws[("C"+ str(7+b+c))] = "*" + str(other[5])
                            ws[("D"+ str(7+b+c))] = other[3]
                            Bank_income += other[3]
                    else:
                        ws[("C"+ str(7+b+c))] = str(other[5])
                        ws[("D"+ str(7+b+c))] = other[3]  
                        cash_income += other[3]
                    area_sum += other[3]
                    ws[("C"+ str(36))] = '소 계'
                    ws[("D"+ str(36))] = area_sum
                    c += 1 
                elif (len(seonkyo_list)-29)+len(area_list)+len(other_list) < 56 and a > 29:
                    if other[4] == '통장예입':
                        ws[("C"+ str(5+a2+b+c))] = "*" + other[5]
                        ws[("D"+ str(5+a2+b+c))] = other[3]
                        Bank_income += other[3]
                    else:
                        ws[("C"+ str(5+a2+b+c))] = other[5]
                        ws[("D"+ str(5+a2+b+c))] = other[3]
                        cash_income += other[3]
                    area_sum += other[3]
                    ws[("C"+ str(36))] = '소 계'
                    ws[("D"+ str(36))] = area_sum
                    c += 1      
                    
            #  지출내역 인쇄
            cos1 = 0; today_sun_cost_sum = 0
            today_sun_cost = today_mission_cost_list(Y1,M1,W1)
            ws[("E"+ str(5))] = "지 출 내 역"
            ws[("E"+ str(6))] = "내 역" ; ws[("G"+ str(6))] = "금 액" ;
            for sun_row in today_sun_cost:
                # if sun_row[3] == '기관선교' or sun_row[3] == '미지립교회' or sun_row[3] == '해외선교사':
                ws[("E"+ str(7+cos1))] = str(sun_row[0])
                # else:
                #     ws[("E"+ str(7+cos1))] = str(sun_row[0])
                ws[("G"+ str(7+cos1))] = sun_row[1]
                today_sun_cost_sum += sun_row[1]
                cos1 += 1
            
            ws[("G"+ str(36))] = today_sun_cost_sum

            ws["E36"]='소  계'
            ws["A37"]='현금수입'
            ws["B37"]= cash_income
            ws["C37"]='통장예입 계'
            ws["D37"]= Bank_income
            ws["A38"] = "지난주 잔액"
            ws["B38"] = past_sun_income_amount - past_sun_cost_amount
            ws["E37"]='금주 선교헌금 계'
            ws["G37"] = Bank_income + cash_income
            ws["C38"] = "금주지출총액"
            ws["D38"] = today_sun_cost_sum
            ws["E38"] = "금주 마감잔액"
            ws["G38"] = (past_sun_income_amount - past_sun_cost_amount) + (Bank_income + cash_income) - today_sun_cost_sum
            
            bogoseo.save('./excel_view/선교회계재정보고.xlsx')
            bogoseo.close()
            self.week_bogo_confirm.setText("{}년도 {}월 {}주 선교회계 재정보고서가 생성되었습니다.".format(Y1,M1,W1))
            subprocess.Popen(["start", "excel.exe", os.path.abspath("./excel_view/선교회계재정보고.xlsx")], shell=True)

        except ValueError:
            QMessageBox.about(self,'입력오류','입력을 확인하세요 !!!')

        except PermissionError:
            QMessageBox.about(self,'파일열기','파일이 열려 있습니다.')
